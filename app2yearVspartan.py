import streamlit as st
import io
import time
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from nselib import capital_market
except ImportError:
    st.error("nselib is not installed. Check your requirements.txt")
    st.stop()

# --- EXCEL FORMATTING ---
def format_excel_sheet(ws, df_sheet, title_text, default_days_or_info):
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    total_font = Font(name=font_family, size=11, bold=True, color="1B365D")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    buy_font = Font(name=font_family, size=10, bold=True, color="006100")
    sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sell_font = Font(name=font_family, size=10, bold=True, color="9C0006")
    
    thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                         top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
    double_bottom_border = Border(top=Side(style='thin', color='1B365D'), bottom=Side(style='double', color='1B365D'))

    ws.views.sheetView[0].showGridLines = True
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{title_text} ({default_days_or_info})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 10
    
    headers = ["Date", "Symbol", "Company Name", "Client Name", "Deal Type", "Quantity", "Price (₹)", "Trade Value (₹)", "Remarks"]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5] else "left" if col_idx in [3, 4, 9] else "right", vertical="center")
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 25
    
    row_idx = 4
    for _, row in df_sheet.iterrows():
        ws.row_dimensions[row_idx].height = 20
        vals = [
            row.get('Date', ''), row.get('Symbol', ''), row.get('SecurityName', ''),
            row.get('ClientName', ''), row.get('Buy/Sell', ''), row.get('QuantityTraded', 0),
            row.get('TradePrice/Wght.Avg.Price', 0.0), row.get('TradeValue_INR', 0.0), row.get('Remarks', '-')
        ]
        
        is_buy = (str(row.get('Buy/Sell', '')).strip().upper() == "BUY")
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            
            if col_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=font_family, size=10, bold=True, color="333333")
            elif col_idx in [3, 4, 9]: cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = buy_fill if is_buy else sell_fill
                cell.font = buy_font if is_buy else sell_font
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##,##0'
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##,##,##0.00'
        row_idx += 1
        
    ws.row_dimensions[row_idx].height = 24
    total_label_cell = ws.cell(row=row_idx, column=1)
    total_label_cell.value = "Total"
    total_label_cell.font = total_font
    total_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    total_label_cell.border = double_bottom_border
    
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    for c in range(2, 6): ws.cell(row=row_idx, column=c).border = double_bottom_border
        
    total_qty_cell = ws.cell(row=row_idx, column=6)
    total_qty_cell.value = f"=SUM(F4:F{row_idx-1})"
    total_qty_cell.font = total_font
    total_qty_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_qty_cell.number_format = '#,##,##0'
    total_qty_cell.border = double_bottom_border
    
    avg_price_cell = ws.cell(row=row_idx, column=7)
    avg_price_cell.value = f"=H{row_idx}/F{row_idx}"
    avg_price_cell.font = total_font
    avg_price_cell.alignment = Alignment(horizontal="right", vertical="center")
    avg_price_cell.number_format = '#,##0.00'
    avg_price_cell.border = double_bottom_border
    
    total_val_cell = ws.cell(row=row_idx, column=8)
    total_val_cell.value = f"=SUM(H4:H{row_idx-1})"
    total_val_cell.font = total_font
    total_val_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_val_cell.number_format = '#,##,##,##0.00'
    total_val_cell.border = double_bottom_border
    
    ws.cell(row=row_idx, column=9).border = double_bottom_border
    
    ws.column_dimensions['A'].width = 13  
    ws.column_dimensions['B'].width = 14  
    ws.column_dimensions['C'].width = 30  
    ws.column_dimensions['D'].width = 45  
    ws.column_dimensions['E'].width = 13  
    ws.column_dimensions['F'].width = 18  
    ws.column_dimensions['G'].width = 14  
    ws.column_dimensions['H'].width = 22  
    ws.column_dimensions['I'].width = 15  

# --- DATA PROCESSING ---
def process_df(df):
    if df.empty: return df
    df.columns = [c.strip() for c in df.columns]
    
    def parse_date(d_str):
        d_str = str(d_str).strip()
        for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try: return datetime.strptime(d_str, fmt)
            except ValueError: continue
        return d_str
        
    df['ParsedDate'] = df['Date'].apply(parse_date)
    
    if pd.api.types.is_datetime64_any_dtype(df['ParsedDate']):
        df = df.sort_values(by=['ParsedDate', 'Symbol'], ascending=[False, True])
        df['Date'] = df['ParsedDate'].dt.strftime('%d-%b-%y')
    else:
        df = df.sort_values(by=['Date', 'Symbol'], ascending=[False, True])
        
    df = df.drop(columns=['ParsedDate'])
    df['QuantityTraded'] = pd.to_numeric(df['QuantityTraded'].astype(str).str.replace(',', ''), errors='coerce').fillna(0).astype(int)
    df['TradePrice/Wght.Avg.Price'] = pd.to_numeric(df['TradePrice/Wght.Avg.Price'].astype(str).str.replace(',', ''), errors='coerce').fillna(0.0)
    df['TradeValue_INR'] = df['QuantityTraded'] * df['TradePrice/Wght.Avg.Price']
    df['Buy/Sell'] = df['Buy/Sell'].astype(str).str.strip().str.upper()
    
    return df

def get_excel_buffer(df_bulk, df_block, title_prefix, sub_title):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) 
    
    if not df_bulk.empty:
        ws_bulk = wb.create_sheet("Bulk Deals")
        format_excel_sheet(ws_bulk, df_bulk, f"{title_prefix} Bulk Deals", sub_title)
        
    if not df_block.empty:
        ws_block = wb.create_sheet("Block Deals")
        format_excel_sheet(ws_block, df_block, f"{title_prefix} Block Deals", sub_title)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- STREAMLIT UI ---
st.set_page_config(page_title="NSE Tracker", layout="wide")
st.title("📊 Client 2-Year Historical Tracker")

client_name = st.text_input("Enter client name keyword:", value="VSPARTANS").strip()

if st.button("Fetch Data", type="primary"):
    if not client_name:
        st.error("Please enter a client name.")
        st.stop()
        
    today = datetime.now()
    all_bulk, all_block = [], []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # 2 Year Loop
    for i in range(2):
        start_days = i * 365
        end_days = (i + 1) * 365
        from_dt = today - timedelta(days=end_days)
        to_dt = today - timedelta(days=start_days)
        from_str = from_dt.strftime("%d-%m-%Y")
        to_str = to_dt.strftime("%d-%m-%Y")
        
        status_text.info(f"📡 Fetching data for Year {i+1} ({from_str} to {to_str}). Please wait...")
        
        try:
            df_b = capital_market.bulk_deal_data(from_date=from_str, to_date=to_str)
            if not df_b.empty:
                df_b.columns = [c.strip() for c in df_b.columns]
                filtered = df_b[df_b['ClientName'].astype(str).str.contains(client_name, case=False, na=False)]
                if not filtered.empty: all_bulk.append(filtered)
        except Exception as e:
            st.warning(f"Bulk data warning for {from_str}: {e}")
            
        time.sleep(1) # Reduced to 1 second to prevent Streamlit Cloud timeout
        
        try:
            df_bl = capital_market.block_deals_data(from_date=from_str, to_date=to_str)
            if not df_bl.empty:
                df_bl.columns = [c.strip() for c in df_bl.columns]
                filtered = df_bl[df_bl['ClientName'].astype(str).str.contains(client_name, case=False, na=False)]
                if not filtered.empty: all_block.append(filtered)
        except Exception as e:
            st.warning(f"Block data warning for {from_str}: {e}")
            
        progress_bar.progress((i + 1) * 50)
        time.sleep(1) 
        
    status_text.empty()
    progress_bar.empty()
    
    df_bulk_merged = pd.concat(all_bulk, ignore_index=True) if all_bulk else pd.DataFrame()
    df_block_merged = pd.concat(all_block, ignore_index=True) if all_block else pd.DataFrame()
    
    if df_bulk_merged.empty and df_block_merged.empty:
        st.error(f"No records found for client: '{client_name}'.")
    else:
        df_bulk_merged = process_df(df_bulk_merged)
        df_block_merged = process_df(df_block_merged)
        
        total_trades = len(df_bulk_merged) + len(df_block_merged)
        st.success(f"✅ Success! Found {total_trades} trades for '{client_name.upper()}'")
        
        if not df_bulk_merged.empty:
            st.subheader("Bulk Deals")
            st.dataframe(df_bulk_merged, use_container_width=True)
            
        if not df_block_merged.empty:
            st.subheader("Block Deals")
            st.dataframe(df_block_merged, use_container_width=True)
        
        sanitized_name = "".join([c for c in client_name if c.isalnum() or c in (' ', '_')]).strip().replace(' ', '_')
        file_name = f"{sanitized_name}_2_Years.xlsx"
        excel_buffer = get_excel_buffer(df_bulk_merged, df_block_merged, f"Client: {client_name.upper()}", "2 Year History")
        
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_buffer,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
