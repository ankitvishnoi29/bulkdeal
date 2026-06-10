import os
import sys
import time
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ensure nselib is installed
try:
    from nselib import capital_market
except ImportError:
    print("nselib is not installed. Please install it using: pip install nselib")
    sys.exit(1)

def format_excel_sheet(ws, df_sheet, title_text, default_days_or_info):
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    total_font = Font(name=font_family, size=11, bold=True, color="1B365D")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Premium Excel standard green/red for BUY/SELL
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    buy_font = Font(name=font_family, size=10, bold=True, color="006100")
    
    sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sell_font = Font(name=font_family, size=10, bold=True, color="9C0006")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='1B365D'),
        bottom=Side(style='double', color='1B365D')
    )

    ws.views.sheetView[0].showGridLines = True
    
    # Title Row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{title_text} ({default_days_or_info})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 10
    
    # Headers
    headers = [
        "Date", "Symbol", "Company Name", "Client Name", 
        "Deal Type", "Quantity", "Price (₹)", "Trade Value (₹)", "Remarks"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5] else "left" if col_idx in [3, 4, 9] else "right", vertical="center")
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 25
    
    # Write Rows
    row_idx = 4
    for _, row in df_sheet.iterrows():
        ws.row_dimensions[row_idx].height = 20
        
        vals = [
            row.get('Date', ''),
            row.get('Symbol', ''),
            row.get('SecurityName', ''),
            row.get('ClientName', ''),
            row.get('Buy/Sell', ''),
            row.get('QuantityTraded', 0),
            row.get('TradePrice/Wght.Avg.Price', 0.0),
            row.get('TradeValue_INR', 0.0),
            row.get('Remarks', '-')
        ]
        
        is_buy = (str(row.get('Buy/Sell', '')).strip().upper() == "BUY")
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=font_family, size=10, bold=True, color="333333")
            elif col_idx in [3, 4, 9]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_buy:
                    cell.fill = buy_fill
                    cell.font = buy_font
                else:
                    cell.fill = sell_fill
                    cell.font = sell_font
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##,##0'
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##,##,##0.00'
                
        row_idx += 1
        
    # Total Row
    ws.row_dimensions[row_idx].height = 24
    total_label_cell = ws.cell(row=row_idx, column=1)
    total_label_cell.value = "Total"
    total_label_cell.font = total_font
    total_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    total_label_cell.border = double_bottom_border
    
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    for c in range(2, 6):
        ws.cell(row=row_idx, column=c).border = double_bottom_border
        
    # Total Qty
    total_qty_cell = ws.cell(row=row_idx, column=6)
    total_qty_cell.value = f"=SUM(F4:F{row_idx-1})"
    total_qty_cell.font = total_font
    total_qty_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_qty_cell.number_format = '#,##,##0'
    total_qty_cell.border = double_bottom_border
    
    # Weighted Average Price
    avg_price_cell = ws.cell(row=row_idx, column=7)
    avg_price_cell.value = f"=H{row_idx}/F{row_idx}"
    avg_price_cell.font = total_font
    avg_price_cell.alignment = Alignment(horizontal="right", vertical="center")
    avg_price_cell.number_format = '#,##0.00'
    avg_price_cell.border = double_bottom_border
    
    # Total Trade Value
    total_val_cell = ws.cell(row=row_idx, column=8)
    total_val_cell.value = f"=SUM(H4:H{row_idx-1})"
    total_val_cell.font = total_font
    total_val_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_val_cell.number_format = '#,##,##,##0.00'
    total_val_cell.border = double_bottom_border
    
    # Border for remarks cell in total row
    ws.cell(row=row_idx, column=9).border = double_bottom_border
    
    # Adjust Columns auto-fit
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and ('#,##' in cell.number_format):
                max_len = max(max_len, len(val_str) + 6)
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Manual Adjustments for key columns
    ws.column_dimensions['A'].width = 13  # Date
    ws.column_dimensions['B'].width = 14  # Symbol
    ws.column_dimensions['C'].width = 30  # Company Name
    ws.column_dimensions['D'].width = 45  # Client Name
    ws.column_dimensions['E'].width = 13  # Deal Type
    ws.column_dimensions['F'].width = 18  # Quantity
    ws.column_dimensions['G'].width = 14  # Price
    ws.column_dimensions['H'].width = 22  # Value
    ws.column_dimensions['I'].width = 15  # Remarks

def process_df(df):
    if df.empty:
        return df
    
    # Clean Column names
    df.columns = [c.strip() for c in df.columns]
    
    # Date formatting
    def parse_date(d_str):
        d_str = str(d_str).strip()
        for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(d_str, fmt)
            except ValueError:
                continue
        return d_str
        
    df['ParsedDate'] = df['Date'].apply(parse_date)
    
    # Sort
    if pd.api.types.is_datetime64_any_dtype(df['ParsedDate']):
        df = df.sort_values(by=['ParsedDate', 'Symbol'], ascending=[False, True])
        df['Date'] = df['ParsedDate'].dt.strftime('%d-%b-%y')
    else:
        df = df.sort_values(by=['Date', 'Symbol'], ascending=[False, True])
        
    df = df.drop(columns=['ParsedDate'])
    
    # Format and clean numeric values
    df['QuantityTraded'] = pd.to_numeric(df['QuantityTraded'].astype(str).str.replace(',', ''), errors='coerce').fillna(0).astype(int)
    df['TradePrice/Wght.Avg.Price'] = pd.to_numeric(df['TradePrice/Wght.Avg.Price'].astype(str).str.replace(',', ''), errors='coerce').fillna(0.0)
    df['TradeValue_INR'] = df['QuantityTraded'] * df['TradePrice/Wght.Avg.Price']
    df['Buy/Sell'] = df['Buy/Sell'].astype(str).str.strip().str.upper()
    
    return df

def run_option_1():
    print("\n--- OPTION 1: DAILY MARKET DEALS ---")
    days_input = input("Enter number of days of market data to fetch [Default: 30]: ").strip()
    days = 30
    if days_input.isdigit():
        days = int(days_input)
        
    today = datetime.now()
    from_date = (today - timedelta(days=days)).strftime("%d-%m-%Y")
    to_date = today.strftime("%d-%m-%Y")
    
    print(f"\nFetching data from {from_date} to {to_date}...")
    
    # Fetch Bulk
    print("Fetching Bulk Deals...")
    try:
        df_bulk = capital_market.bulk_deal_data(from_date=from_date, to_date=to_date)
        df_bulk = process_df(df_bulk)
        print(f"-> Successfully fetched {len(df_bulk)} Bulk Deals.")
    except Exception as e:
        print(f"-> Error: {e}")
        df_bulk = pd.DataFrame()
        
    # Fetch Block
    print("Fetching Block Deals...")
    try:
        df_block = capital_market.block_deals_data(from_date=from_date, to_date=to_date)
        df_block = process_df(df_block)
        print(f"-> Successfully fetched {len(df_block)} Block Deals.")
    except Exception as e:
        print(f"-> Error: {e}")
        df_block = pd.DataFrame()
        
    if df_bulk.empty and df_block.empty:
        print("No data fetched from NSE. Exiting...")
        return
        
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    file_name = f"NSE_Market_Deals_Last_{days}_Days.xlsx"
    
    if not df_bulk.empty:
        ws_bulk = wb.create_sheet("Bulk Deals")
        format_excel_sheet(ws_bulk, df_bulk, "NSE Bulk Deals", f"Last {days} Days as of {today.strftime('%d-%b-%Y')}")
        
    if not df_block.empty:
        ws_block = wb.create_sheet("Block Deals")
        format_excel_sheet(ws_block, df_block, "NSE Block Deals", f"Last {days} Days as of {today.strftime('%d-%b-%Y')}")
        
    wb.save(file_name)
    print(f"\n[SUCCESS] Excel file created: '{file_name}'")
    print("Aap is file ko Google Drive par upload karke direct use kar sakte hain!")

def run_option_2():
    print("\n--- OPTION 2: CLIENT 10-YEAR HISTORICAL TRACKER ---")
    client_name = input("Enter client name keyword [Default: VSPARTANS]: ").strip()
    if not client_name:
        client_name = "VSPARTANS"
        
    print(f"\nSearching NSE large deal database (2016-2026) for client: '{client_name.upper()}'...")
    
    today = datetime.now()
    all_bulk = []
    all_block = []
    
    # Loop over 10 years in chunks of 1 year to avoid server timeouts
    for i in range(10):
        start_days = i * 365
        end_days = (i + 1) * 365
        
        from_dt = today - timedelta(days=end_days)
        to_dt = today - timedelta(days=start_days)
        
        from_str = from_dt.strftime("%d-%m-%Y")
        to_str = to_dt.strftime("%d-%m-%Y")
        
        print(f"-> Searching Year {i+1}/10 ({from_str} to {to_str})...")
        
        # Search bulk deals
        try:
            df_b = capital_market.bulk_deal_data(from_date=from_str, to_date=to_str)
            if not df_b.empty:
                df_b.columns = [c.strip() for c in df_b.columns]
                filtered = df_b[df_b['ClientName'].astype(str).str.contains(client_name, case=False, na=False)]
                if not filtered.empty:
                    print(f"   * Found {len(filtered)} Bulk Deals!")
                    all_bulk.append(filtered)
        except Exception as e:
            print(f"   * Error fetching bulk: {e}")
            
        # Search block deals
        try:
            df_bl = capital_market.block_deals_data(from_date=from_str, to_date=to_str)
            if not df_bl.empty:
                df_bl.columns = [c.strip() for c in df_bl.columns]
                filtered = df_bl[df_bl['ClientName'].astype(str).str.contains(client_name, case=False, na=False)]
                if not filtered.empty:
                    print(f"   * Found {len(filtered)} Block Deals!")
                    all_block.append(filtered)
        except Exception as e:
            print(f"   * Error fetching block: {e}")
            
        time.sleep(0.5) # Sleep to avoid spamming
        
    df_bulk_merged = pd.concat(all_bulk, ignore_index=True) if all_bulk else pd.DataFrame()
    df_block_merged = pd.concat(all_block, ignore_index=True) if all_block else pd.DataFrame()
    
    if df_bulk_merged.empty and df_block_merged.empty:
        print(f"\n[INFO] Client '{client_name}' ke liye koi data nahi mila.")
        return
        
    df_bulk_merged = process_df(df_bulk_merged)
    df_block_merged = process_df(df_block_merged)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    sanitized_name = "".join([c for c in client_name if c.isalnum() or c in (' ', '_')]).strip().replace(' ', '_')
    file_name = f"Client_History_{sanitized_name}.xlsx"
    
    if not df_bulk_merged.empty:
        ws_b = wb.create_sheet("Bulk Deals")
        format_excel_sheet(ws_b, df_bulk_merged, f"Client: {client_name.upper()} (Bulk)", "10 Year History")
        
    if not df_block_merged.empty:
        ws_bl = wb.create_sheet("Block Deals")
        format_excel_sheet(ws_bl, df_block_merged, f"Client: {client_name.upper()} (Block)", "10 Year History")
        
    wb.save(file_name)
    print(f"\n[SUCCESS] Excel file created: '{file_name}'")
    print(f"Client '{client_name}' ke total {len(df_bulk_merged) + len(df_block_merged)} trades save ho chuke hain!")
    print("Aap is file ko Google Drive par upload karke direct Google Sheets mein view kar sakte hain.")

def main():
    print("=" * 60)
    print("       NSE BULK & BLOCK DEALS INTERACTIVE TRACKER")
    print("=" * 60)
    print("1. Fetch Daily Market Deals (Specify days, e.g. 30 days)")
    print("2. Track Specific Client (10-Year Trade History)")
    print("-" * 60)
    
    choice = input("Enter choice (1 or 2): ").strip()
    if choice == '1':
        run_option_1()
    elif choice == '2':
        run_option_2()
    else:
        print("Invalid choice. Please run the script again and select 1 or 2.")

if __name__ == "__main__":
    main()
